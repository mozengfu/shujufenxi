"""报表配置数据模型和生成引擎"""
import json
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field
import pandas as pd
from matplotlib.figure import Figure

from core.reporter import WordReporter
from core.exporter import ExcelExporter
from core.analyzer import DataAnalyzer


@dataclass
class ReportSection:
    """报表章节配置"""
    section_type: str  # 'title', 'stats', 'quality', 'comparison', 'chart', 'data_table', 'text'
    title: str = ''
    config: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {'section_type': self.section_type, 'title': self.title, 'config': self.config}

    @classmethod
    def from_dict(cls, d: dict) -> 'ReportSection':
        return cls(section_type=d['section_type'], title=d.get('title', ''), config=d.get('config', {}))


@dataclass
class ReportConfig:
    """完整报表配置"""
    sections: List[ReportSection] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {'sections': [s.to_dict() for s in self.sections]}

    @classmethod
    def from_dict(cls, d: dict) -> 'ReportConfig':
        sections = [ReportSection.from_dict(s) for s in d.get('sections', [])]
        return cls(sections=sections)

    def to_json(self) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=2)

    @classmethod
    def from_json(cls, text: str) -> 'ReportConfig':
        return cls.from_dict(json.loads(text))


class ReportGenerator:
    """报表生成引擎，根据 ReportConfig 生成 Word/Excel 报告"""

    def __init__(self, df: pd.DataFrame, analyzer: DataAnalyzer):
        self.df = df
        self.analyzer = analyzer

    def generate_word(self, config: ReportConfig, reporter: WordReporter,
                      chart_figures: Optional[Dict[str, Figure]] = None) -> WordReporter:
        """根据配置生成 Word 文档"""
        reporter.create_document()
        chart_figures = chart_figures or {}

        for section in config.sections:
            self._render_word_section(section, reporter, chart_figures)

        return reporter

    def _render_word_section(self, section: ReportSection, reporter: WordReporter,
                              chart_figures: Dict[str, Figure]):
        """渲染单个章节到 Word"""
        stype = section.section_type
        title = section.title or self._default_title(stype)

        if stype == 'title':
            reporter.add_title(title, level=0)

        elif stype == 'text':
            reporter.add_title(title, level=1)
            text = section.config.get('text', '')
            reporter.add_paragraph(text)

        elif stype == 'stats':
            reporter.add_title(title, level=1)
            cols = section.config.get('columns')
            if cols:
                stats = self.analyzer.describe_cols(self.df, cols)
            else:
                stats = self.analyzer.descriptive_stats(self.df)
            if not stats.empty:
                reporter.add_table_from_df(stats.round(2))

        elif stype == 'quality':
            reporter.add_title(title, level=1)
            report = self.analyzer.full_quality_report(self.df)
            # 按需筛选质量报告内容
            reporter.add_quality_report(report)

        elif stype == 'comparison':
            comparison = section.config.get('comparison_data')
            if comparison:
                reporter.add_comparison_report(comparison)

        elif stype == 'chart':
            fig_key = section.config.get('figure_key', '')
            fig = chart_figures.get(fig_key)
            if fig is not None:
                reporter.add_chart(fig, title=title)

        elif stype == 'data_table':
            reporter.add_title(title, level=1)
            cols = section.config.get('columns')
            max_rows = section.config.get('max_rows', 50)
            if cols:
                subset = self.df[cols].head(max_rows)
            else:
                subset = self.df.head(max_rows)
            reporter.add_table_from_df(subset)

    def generate_excel(self, config: ReportConfig, exporter: ExcelExporter, file_path: str):
        """根据配置生成 Excel 报告（每个章节一个工作表）"""
        from openpyxl.styles import Font, PatternFill

        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for i, section in enumerate(config.sections):
            stype = section.section_type
            sheet_name = section.title[:31] or f'章节{i+1}'

            if stype == 'stats':
                cols = section.config.get('columns')
                if cols:
                    data = self.analyzer.describe_cols(self.df, cols)
                else:
                    data = self.analyzer.descriptive_stats(self.df)
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=True)

            elif stype == 'data_table':
                cols = section.config.get('columns')
                max_rows = section.config.get('max_rows', 1000)
                if cols:
                    data = self.df[cols].head(max_rows)
                else:
                    data = self.df.head(max_rows)
                if not data.empty:
                    data.to_excel(writer, sheet_name=sheet_name, index=False)

            elif stype == 'quality':
                report = self.analyzer.full_quality_report(self.df)
                exporter.export_quality_report(report, file_path)

            elif stype == 'text':
                pd.DataFrame({'内容': [section.config.get('text', '')]}).to_excel(
                    writer, sheet_name=sheet_name, index=False)

        writer.close()

        # 格式美化
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
        wb.save(file_path)

    def _default_title(self, section_type: str) -> str:
        """为章节类型生成默认标题"""
        titles = {
            'title': '数据分析报告',
            'text': '说明',
            'stats': '描述性统计',
            'quality': '数据质量报告',
            'comparison': '对比分析',
            'chart': '分析图表',
            'data_table': '数据明细',
        }
        return titles.get(section_type, '未命名章节')
