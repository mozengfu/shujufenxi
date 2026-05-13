"""Excel 导出模块"""
import pandas as pd
from pathlib import Path
from typing import Dict, Any, List, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import unicodedata
from core.analyzer import make_total_row, is_percent_col, _is_numeric


class ExcelExporter:
    """Excel 导出器"""

    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True)
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.data_align = Alignment(vertical="center", wrap_text=True)
        self.total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        self.total_font = Font(bold=True)
        self.total_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.min_row_height = 15  # 最小行高
        self.max_row_height = 60   # 最大行高
        self.base_col_width = 1.5  # 基础列宽系数
        # 边框样式
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.header_border = Border(
            left=Side(style='thin', color='FFFFFF'),
            right=Side(style='thin', color='FFFFFF'),
            top=Side(style='thin', color='FFFFFF'),
            bottom=Side(style='thin', color='FFFFFF')
        )

    def export_dataframe(self, df: pd.DataFrame, file_path: str, sheet_name: str = 'Sheet1') -> None:
        """导出 DataFrame 到 Excel"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')

    def export_with_format(self, df: pd.DataFrame, file_path: str, sheet_name: str = 'Sheet1',
                           agg_items: Optional[List[Dict]] = None) -> None:
        """带格式导出 DataFrame 到 Excel（含合计行）"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        # 追加合计行
        export_df = self._add_total_row(df, agg_items)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 应用格式
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # 应用完整格式（边框、样式、数字格式）
        self._apply_table_format(ws, export_df, has_total_row=True)

        # 冻结首行和自动筛选
        self._freeze_header_and_filter(ws, export_df)

        # 打印区域和页面设置
        self._setup_print_settings(ws, export_df)

        # 列宽和行高自适应
        self._adjust_column_widths(ws, export_df)
        self._adjust_row_heights(ws, export_df)

        wb.save(file_path)

    def _apply_table_format(self, ws, df: pd.DataFrame, has_total_row: bool = False) -> None:
        """应用表格格式：边框、样式、数字格式"""
        total_row_idx = len(df) + 1 if has_total_row else len(df)

        # 表头格式（第1行）
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.header_border

        # 数据行格式
        for row_idx in range(2, total_row_idx + 1):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = self.data_align
                cell.border = self.thin_border

                # 数字格式处理
                if _is_numeric(df[col_name]):
                    # 检查是否是百分比列
                    if is_percent_col(col_name):
                        # 已经是字符串格式，保持原样
                        pass
                    else:
                        # 普通数值：千分位格式
                        cell.number_format = '#,##0.00'
                elif isinstance(cell.value, str) and '%' in cell.value:
                    # 尝试转换为数字并应用百分比格式
                    try:
                        val = float(cell.value.replace('%', ''))
                        cell.value = val / 100
                        cell.number_format = '0.00%'
                    except (ValueError, TypeError):
                        pass

        # 合计行格式
        if has_total_row:
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=total_row_idx + 1, column=col_idx)
                cell.fill = self.total_fill
                cell.font = self.total_font
                cell.alignment = self.total_align
                cell.border = self.thin_border

    def _freeze_header_and_filter(self, ws, df: pd.DataFrame) -> None:
        """冻结首行并添加自动筛选"""
        # 冻结首行
        ws.freeze_panes = 'A2'

        # 添加自动筛选
        if len(df) > 0:
            last_col = get_column_letter(len(df.columns))
            ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    def _setup_print_settings(self, ws, df: pd.DataFrame) -> None:
        """设置打印区域和页面设置"""
        # 页面设置
        ws.page_setup.orientation = 'landscape' if len(df.columns) > 6 else 'portrait'
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # 页边距
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

        # 打印标题行
        ws.print_title_rows = '1:1'

        # 打印网格线
        ws.print_gridlines = False

    def _get_text_width(self, text: str) -> float:
        """计算文本宽度，中文字符按2个宽度计算"""
        width = 0
        for char in str(text):
            if unicodedata.east_asian_width(char) in ('F', 'W', 'A'):
                width += 2
            else:
                width += 1
        return width

    def _adjust_column_widths(self, ws, df: pd.DataFrame) -> None:
        """自适应列宽"""
        for col_idx, column in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            col_data = df[column].astype(str)

            # 计算表头宽度
            header_width = self._get_text_width(str(column))
            max_len = header_width

            # 计算数据最大宽度（采样前100行避免大数据量过慢）
            if len(df) > 0:
                sample_data = col_data.head(100)
                for val in sample_data:
                    val_width = self._get_text_width(val)
                    max_len = max(max_len, val_width)

            # 设置列宽（加边距，最小8最大50）
            adjusted_width = min(max(max_len * self.base_col_width + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _adjust_row_heights(self, ws, df: pd.DataFrame) -> None:
        """自适应行高"""
        # 表头行高
        header_lines = 1
        for col in df.columns:
            lines = str(col).count('\n') + 1
            header_lines = max(header_lines, lines)
        ws.row_dimensions[1].height = min(header_lines * 15, self.max_row_height)

        # 数据行高
        for row_idx in range(2, len(df) + 2):  # +2 包含合计行
            max_lines = 1
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    # 根据列宽估算需要的行数
                    col_letter = get_column_letter(col_idx)
                    col_width = ws.column_dimensions[col_letter].width
                    text_len = self._get_text_width(str(cell.value))
                    if col_width > 0:
                        lines_needed = max(1, int(text_len / col_width) + 1)
                        max_lines = max(max_lines, lines_needed)

            # 设置行高（每行约15点）
            height = min(max(max_lines * 15, self.min_row_height), self.max_row_height)
            ws.row_dimensions[row_idx].height = height

    def _apply_format_to_worksheet(self, ws, header_row: int = 1) -> None:
        """为 worksheet 应用格式（边框、数字格式）"""
        # 表头格式
        for cell in ws[header_row]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.header_border

        # 数据行格式
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = self.data_align
                cell.border = self.thin_border

                # 尝试应用数字格式
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif isinstance(cell.value, str):
                    # 检查是否是百分比
                    if '%' in cell.value:
                        try:
                            val = float(cell.value.replace('%', ''))
                            cell.value = val / 100
                            cell.number_format = '0.00%'
                        except (ValueError, TypeError):
                            pass

    def _adjust_column_widths_for_worksheet(self, ws) -> None:
        """为已存在的 worksheet 自适应列宽"""
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0

            # 遍历该列所有单元格
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    val_width = self._get_text_width(str(cell.value))
                    max_len = max(max_len, val_width)

            # 设置列宽
            adjusted_width = min(max(max_len * self.base_col_width + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _adjust_row_heights_for_worksheet(self, ws) -> None:
        """为已存在的 worksheet 自适应行高"""
        for row_idx in range(1, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    col_letter = get_column_letter(col_idx)
                    col_width = ws.column_dimensions[col_letter].width or 8
                    text_len = self._get_text_width(str(cell.value))
                    if col_width > 0:
                        lines_needed = max(1, int(text_len / col_width) + 1)
                        max_lines = max(max_lines, lines_needed)

            height = min(max(max_lines * 15, self.min_row_height), self.max_row_height)
            ws.row_dimensions[row_idx].height = height

    def _add_total_row(self, df: pd.DataFrame, agg_items: Optional[List[Dict]] = None) -> pd.DataFrame:
        """给 DataFrame 追加合计行，占比列格式化为百分字符串"""
        if df is None or df.empty:
            return df
        export_df = df.copy()

        # 用共享逻辑计算合计行（含占比列的正确合计值）
        total_row = make_total_row(export_df, agg_items)

        # 占比列格式化为百分字符串（数据行 + 合计行）
        for col in export_df.columns:
            if is_percent_col(col) and _is_numeric(export_df[col]):
                export_df[col] = export_df[col].apply(
                    lambda x: f'{x:.2f}%' if pd.notna(x) else ''
                )
            # 合计行中的占比列也格式化
            if is_percent_col(col):
                val = total_row[col]
                if isinstance(val, (int, float)) and pd.notna(val):
                    total_row[col] = f'{val:.2f}%'
                elif val == '-':
                    total_row[col] = '-'

        return pd.concat([export_df.reset_index(drop=True), total_row.to_frame().T], ignore_index=True)

    def export_stats_report(self, stats: pd.DataFrame, file_path: str, title: str = '描述性统计报告') -> None:
        """导出统计报告"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 标题
            pd.DataFrame({'报告': [title]}).to_excel(writer, sheet_name='统计', index=False, startrow=0)

            # 统计数据
            stats.to_excel(writer, sheet_name='统计', index=True, startrow=2)

        # 应用格式
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb['统计']

        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')

        # 应用完整格式
        self._apply_format_to_worksheet(ws, header_row=3)

        # 冻结首行和自动筛选
        ws.freeze_panes = 'A4'
        if ws.max_row > 3:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A3:{last_col}{ws.max_row}"

        # 打印设置
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws.print_title_rows = '3:3'

        # 自适应列宽和行高
        self._adjust_column_widths_for_worksheet(ws)
        self._adjust_row_heights_for_worksheet(ws)

        wb.save(file_path)

    def export_quality_report(self, quality_report: Dict[str, Any], file_path: str) -> None:
        """导出数据质量报告"""
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 缺失值
            if quality_report.get('missing'):
                missing_data = []
                for col, info in quality_report['missing'].items():
                    missing_data.append({
                        '列名': col,
                        '缺失数量': info['count'],
                        '缺失百分比': f"{info['percentage']}%",
                        '行号': ', '.join(map(str, info['rows'][:20]))
                    })
                pd.DataFrame(missing_data).to_excel(writer, sheet_name='缺失值', index=False)

            # 异常值
            if quality_report.get('outliers'):
                outlier_data = []
                for col, info in quality_report['outliers'].items():
                    outlier_data.append({
                        '列名': col,
                        '异常数量': info['count'],
                        '下界': info.get('lower_bound'),
                        '上界': info.get('upper_bound'),
                        '值': ', '.join(map(str, info['values'][:10]))
                    })
                pd.DataFrame(outlier_data).to_excel(writer, sheet_name='异常值', index=False)

            # 重复行
            if quality_report.get('duplicates'):
                dup = quality_report['duplicates']
                pd.DataFrame([{
                    '重复行数': dup['count'],
                    '占比': f"{dup['percentage']}%",
                    '行号': ', '.join(map(str, dup['rows'][:20]))
                }]).to_excel(writer, sheet_name='重复行', index=False)

        # 应用格式
        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 应用完整格式
            self._apply_format_to_worksheet(ws, header_row=1)

            # 冻结首行和自动筛选
            ws.freeze_panes = 'A2'
            if ws.max_row > 1:
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

            # 打印设置
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize = 9
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
            ws.print_title_rows = '1:1'

            # 自适应列宽和行高
            self._adjust_column_widths_for_worksheet(ws)
            self._adjust_row_heights_for_worksheet(ws)

        wb.save(file_path)

    def export_complex_report(self, df: pd.DataFrame, template, file_path: str,
                              sheet_name: str = '报表') -> None:
        """
        导出复杂报表（支持多层表头、合并单元格、条件格式）

        Args:
            df: 数据 DataFrame
            template: ComplexReportTemplate 实例
            file_path: 输出文件路径
            sheet_name: 工作表名称
        """
        from core.complex_report import ComplexReportGenerator

        # 生成报表数据
        generator = ComplexReportGenerator(template)
        report_df = generator.generate(df)

        # 计算合计行
        total_row = generator.calculate_total_row(report_df)
        if total_row is not None:
            report_df = pd.concat([report_df, total_row.to_frame().T], ignore_index=True)

        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        # 获取表头结构
        header_structure = generator.get_header_structure()
        header_levels = len(header_structure)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 写入数据（从表头行之后开始）
            report_df.to_excel(writer, sheet_name=sheet_name, index=False,
                              startrow=header_levels, header=False)

        # 应用格式
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # 1. 写入并合并多层表头
        self._write_multi_headers(ws, header_structure)

        # 2. 应用数据格式
        self._apply_complex_data_format(ws, report_df, header_levels, template)

        # 3. 应用条件格式
        self._apply_conditional_format(ws, report_df, header_levels, template)

        # 4. 冻结表头
        freeze_row = header_levels + 1
        ws.freeze_panes = f'A{freeze_row}'

        # 5. 自动筛选
        last_col = get_column_letter(len(report_df.columns))
        ws.auto_filter.ref = f"A{freeze_row}:{last_col}{ws.max_row}"

        # 6. 打印设置
        ws.page_setup.orientation = 'landscape' if len(report_df.columns) > 8 else 'portrait'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
        ws.print_title_rows = f'1:{header_levels}'

        # 7. 自适应列宽行高
        self._adjust_complex_column_widths(ws, header_structure, report_df)
        self._adjust_complex_row_heights(ws, header_levels, report_df)

        wb.save(file_path)

    def _write_multi_headers(self, ws, header_structure: List[List[Dict]]) -> None:
        """写入多层表头并合并单元格"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        col_idx = 1
        for level_idx, level in enumerate(header_structure):
            for cell_def in level:
                # 写入单元格值
                cell = ws.cell(row=level_idx + 1, column=col_idx, value=cell_def['name'])
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = thin_border

                # 合并单元格
                if cell_def.get('rowspan', 1) > 1 or cell_def.get('colspan', 1) > 1:
                    start_row = level_idx + 1
                    end_row = start_row + cell_def.get('rowspan', 1) - 1
                    start_col = col_idx
                    end_col = col_idx + cell_def.get('colspan', 1) - 1

                    ws.merge_cells(
                        start_row=start_row, start_column=start_col,
                        end_row=end_row, end_column=end_col
                    )

                col_idx += cell_def.get('colspan', 1)

    def _apply_complex_data_format(self, ws, df: pd.DataFrame, header_levels: int,
                                   template) -> None:
        """应用数据格式"""
        data_align = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        total_font = Font(bold=True)

        for row_idx in range(header_levels + 1, ws.max_row + 1):
            is_total_row = (row_idx == ws.max_row and template.total_row.enabled)

            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # 合计行样式
                if is_total_row:
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = data_align

                cell.border = thin_border

                # 数字格式
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'

    def _apply_conditional_format(self, ws, df: pd.DataFrame, header_levels: int,
                                  template) -> None:
        """应用条件格式"""
        for cf in template.conditional_formats:
            # 找到字段对应的列索引
            col_idx = None
            for idx, col in enumerate(df.columns, 1):
                if col == cf.field:
                    col_idx = idx
                    break

            if col_idx is None:
                continue

            # 应用规则
            for rule in cf.rules:
                value = rule.get('value')
                style = rule.get('style', {})

                for row_idx in range(header_levels + 1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value == value:
                        if 'bg_color' in style:
                            cell.fill = PatternFill(
                                start_color=style['bg_color'],
                                end_color=style['bg_color'],
                                fill_type="solid"
                            )
                        if 'font_color' in style:
                            cell.font = Font(color=style['font_color'])

    def _adjust_complex_column_widths(self, ws, header_structure: List[List[Dict]],
                                      df: pd.DataFrame) -> None:
        """调整复杂报表的列宽"""
        # 获取叶子节点（实际数据列）
        leaf_cols = []
        if header_structure:
            for cell in header_structure[-1]:
                leaf_cols.append(cell['name'])

        for col_idx, col_name in enumerate(leaf_cols if leaf_cols else df.columns, 1):
            col_letter = get_column_letter(col_idx)

            # 计算最大宽度
            max_len = len(str(col_name))

            if col_name in df.columns:
                col_data = df[col_name].astype(str)
                for val in col_data.head(100):
                    val_width = self._get_text_width(val)
                    max_len = max(max_len, val_width)

            adjusted_width = min(max(max_len * self.base_col_width + 2, 8), 50)
            ws.column_dimensions[col_letter].width = adjusted_width

    def _adjust_complex_row_heights(self, ws, header_levels: int, df: pd.DataFrame) -> None:
        """调整复杂报表的行高"""
        # 表头行高
        for row_idx in range(1, header_levels + 1):
            ws.row_dimensions[row_idx].height = 30

        # 数据行高
        for row_idx in range(header_levels + 1, ws.max_row + 1):
            max_lines = 1
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    col_letter = get_column_letter(col_idx)
                    col_width = ws.column_dimensions[col_letter].width or 8
                    text_len = self._get_text_width(str(cell.value))
                    if col_width > 0:
                        lines_needed = max(1, int(text_len / col_width) + 1)
                        max_lines = max(max_lines, lines_needed)

            height = min(max(max_lines * 15, self.min_row_height), self.max_row_height)
            ws.row_dimensions[row_idx].height = height

    def export_comparison(self, comparison_result: Dict[str, Any], file_path: str) -> None:
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # 基本对比信息
            comparison_df = pd.DataFrame([
                {'对比项': 'df1行数', '值': comparison_result.get('df1_rows', comparison_result.get('row_diff', 0))},
                {'对比项': 'df2行数', '值': comparison_result.get('df2_rows', '')},
                {'对比项': '结果行数', '值': comparison_result.get('result_rows', '')},
            ])
            comparison_df.to_excel(writer, sheet_name='对比结果', index=False)

            # 列差异
            col_diff = []
            if comparison_result.get('only_in_df1'):
                col_diff.append({'类型': '仅在表1', '列名': ', '.join(comparison_result['only_in_df1'])})
            if comparison_result.get('only_in_df2'):
                col_diff.append({'类型': '仅在表2', '列名': ', '.join(comparison_result['only_in_df2'])})
            if col_diff:
                pd.DataFrame(col_diff).to_excel(writer, sheet_name='列差异', index=False)

        # 应用格式
        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 应用完整格式
            self._apply_format_to_worksheet(ws, header_row=1)

            # 冻结首行和自动筛选
            ws.freeze_panes = 'A2'
            if ws.max_row > 1:
                last_col = get_column_letter(ws.max_column)
                ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

            # 打印设置
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.paperSize = 9
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
            ws.print_title_rows = '1:1'

            # 自适应列宽和行高
            self._adjust_column_widths_for_worksheet(ws)
            self._adjust_row_heights_for_worksheet(ws)

        wb.save(file_path)